from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.http import JsonResponse
from .forms import ProductoForm, MovimientoInventarioForm
from django.shortcuts import redirect, render
from django.forms import inlineformset_factory
from .forms import VentaForm, DetallesVentaForm
from .forms import CompraForm, DetalleCompraForm, BalanceForm
from .models import (
    Venta,
    DetallesVenta,
    Producto,
    Compra,
    DetalleCompra,
    MovimientoInventario,
    Categoria,
    Balance,
    FamiliaProducto,
    UnidadMedida,
)
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import Group
from django.views.generic import FormView
from django.db.models import F
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.contrib import messages
from django.core.files.storage import default_storage
from django.db import transaction, IntegrityError, DatabaseError
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from .utils import calcular_perdidas_devolucion, calcular_balance_mensual
from .serializers import VentaCreateSerializer
from django.views.generic import TemplateView
from django.utils.timezone import now
from django.db.models import Sum
from collections import defaultdict
from decimal import Decimal
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from functools import wraps
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import ensure_csrf_cookie
from django.middleware.csrf import get_token
import json



def index(request):
    productos_bajo_stock = Producto.objects.filter(stock_actual__lt=F('stock_minimo')).count()
    return render(request, 'core/index.html', {'alerta_stock': productos_bajo_stock})

login_decorador = method_decorator(login_required, name='dispatch')


def group_decorator(group: str):
    """Return a method decorator enforcing membership in the given group or admin."""

    def decorator(view_func):
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            user = request.user
            if not user.is_authenticated:
                return redirect("login")
            if not (
                user.is_superuser
                or user.groups.filter(name__in=["admin", group]).exists()
            ):
                return HttpResponseForbidden()
            return view_func(request, *args, **kwargs)

        return _wrapped

    return method_decorator(decorator, name="dispatch")

def inferir_familia(nombre_categoria: str) -> FamiliaProducto:
    nombre = (nombre_categoria or "").lower()
    if "bebida" in nombre:
        clave = FamiliaProducto.Clave.BEBIDAS
    elif "empanad" in nombre:
        clave = FamiliaProducto.Clave.EMPANADAS
    elif "ingred" in nombre or "insumo" in nombre:
        clave = FamiliaProducto.Clave.INGREDIENTES
    else:
        clave = FamiliaProducto.Clave.OTROS
    return FamiliaProducto.objects.get(clave=clave)


def resolver_unidad_media(valor):
    valor_normalizado = str(valor).strip() if valor is not None else ""
    if not valor_normalizado:
        valor_normalizado = "unidad"

    if isinstance(valor, (int, Decimal)):
        unidad = UnidadMedida.objects.filter(id=valor).first()
        if unidad:
            return unidad, None
    elif isinstance(valor, float) and valor.is_integer():
        unidad = UnidadMedida.objects.filter(id=int(valor)).first()
        if unidad:
            return unidad, None

    unidad = UnidadMedida.objects.filter(abreviatura__iexact=valor_normalizado).first()
    if unidad:
        return unidad, None
    unidad = UnidadMedida.objects.filter(nombre__iexact=valor_normalizado).first()
    if unidad:
        return unidad, None

    return None, f'Unidad de medida "{valor_normalizado}" no encontrada.'

def obtener_categoria_importacion(nombre_categoria: str) -> tuple[str, FamiliaProducto]:
    nombre = nombre_categoria or settings.IMPORT_DEFAULT_CATEGORY_NAME
    if nombre == settings.IMPORT_DEFAULT_CATEGORY_NAME:
        familia = FamiliaProducto.objects.get(clave=FamiliaProducto.Clave.OTROS)
    else:
        familia = inferir_familia(nombre)
    return nombre, familia

@ensure_csrf_cookie
def login_view(request):
    """Authenticate and log in a user using username and password."""
    if request.method == "POST":
        if request.content_type == "application/json":
            data = json.loads(request.body.decode())
            username = data.get("username")
            password = data.get("password")
        else:
            username = request.POST.get("username")
            password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.content_type == "application/json":
                return JsonResponse({"success": True})
            if user.is_superuser or user.groups.filter(name="admin").exists():
                return redirect("dashboard")
            return redirect("index")
        error_msg = "Usuario o contraseña incorrectos."
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.content_type == "application/json":
            return JsonResponse({"success": False, "error": error_msg}, status=400)
        messages.error(request, error_msg)
    return render(request, "core/login.html")


@ensure_csrf_cookie
def csrf_token_view(request):
    return JsonResponse({"csrfToken": get_token(request)})
    
    
class ProductoListView(ListView):
    model = Producto
    template_name = 'core/product_list.html'
    context_object_name = 'productos'

@login_decorador
class ProductoCreateView(CreateView):
    model = Producto
    form_class = ProductoForm
    template_name = 'core/producto_form.html'
    success_url = reverse_lazy('producto_list')

    def form_valid(self, form):
        messages.success(self.request, "Producto creado correctamente.")
        return super().form_valid(form)

@login_decorador
class ProductoUpdateView(UpdateView):
    model = Producto
    form_class = ProductoForm
    template_name = 'core/producto_form.html'
    success_url = reverse_lazy('producto_list')

    def form_valid(self, form):
        messages.success(self.request, "Producto actualizado correctamente.")
        return super().form_valid(form)

@login_decorador
class ProductoDeleteView(DeleteView):
    model = Producto
    template_name = 'core/producto_confirm_delete.html'
    success_url = reverse_lazy('producto_list')

    def delete(self, request, *args, **kwargs):
        """Elimina el producto y registra la salida en el inventario."""
        producto = self.get_object()
        stock = producto.stock_actual
        if stock > 0:
            MovimientoInventario.objects.create(
                producto=producto,
                tipo="salida",
                cantidad=stock,
                motivo="Eliminación de producto",
                usuario=request.user if request.user.is_authenticated else None,
                operacion_tipo=MovimientoInventario.OPERACION_ELIMINACION,
            )
            messages.warning(
                request,
                f"Se registró una salida de {stock} por eliminación del producto.",
            )
        messages.success(request, "Producto eliminado correctamente.")
        self.object = producto
        producto.delete()
        return redirect(self.success_url)

    

@group_decorator("ventas")
class VentaCreateView(CreateView):
    model = Venta
    form_class = VentaForm
    template_name = 'core/venta_form.html'
    success_url = reverse_lazy('venta_list')

    def get(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        DetalleFormSet = inlineformset_factory(Venta, DetallesVenta, form=DetallesVentaForm, extra=1, can_delete=True)
        formset = DetalleFormSet()
        return render(request, self.template_name, {'form': form, 'formset': formset})

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        DetalleFormSet = inlineformset_factory(Venta, DetallesVenta, form=DetallesVentaForm, extra=1, can_delete=True)
        formset = DetalleFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            detalles_payload = []
            for detalle in formset:
                if not detalle.cleaned_data or detalle.cleaned_data.get("DELETE"):
                    continue
                producto = detalle.cleaned_data["producto"]
                detalles_payload.append(
                    {
                        "producto": producto.id,
                        "cantidad": detalle.cleaned_data["cantidad"],
                        "precio_unitario": detalle.cleaned_data["precio_unitario"],
                    }
                )

            if not detalles_payload:
                form.add_error(None, "Debe agregar al menos un producto a la venta.")
                return render(
                    request, self.template_name, {"form": form, "formset": formset}
                )

            venta_data = {
                "fecha": form.cleaned_data["fecha"],
                "cliente": form.cleaned_data["cliente"].id,
                "detalles": detalles_payload,
            }
            serializer = VentaCreateSerializer(
                data=venta_data, context={"request": request}
            )
            if serializer.is_valid():
                serializer.save()

                messages.success(request, "Venta registrada correctamente.")
                return redirect(self.success_url)
            
            for errores in serializer.errors.values():
                if isinstance(errores, (list, tuple)):
                    for error in errores:
                        form.add_error(None, error)
                else:
                    form.add_error(None, errores)
        
        return render(request, self.template_name, {'form': form, 'formset': formset})

@group_decorator("ventas")
class VentaListView(ListView):
    model = Venta
    template_name = 'core/venta_list.html'
    context_object_name = 'ventas'
    ordering = ['-fecha']

@group_decorator("compras")
class CompraCreateView(CreateView):
    model = Compra
    form_class = CompraForm
    template_name = 'core/compra_form.html'
    success_url = reverse_lazy('compra_list')

    def get(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        DetalleFormSet = inlineformset_factory(Compra, DetalleCompra, form=DetalleCompraForm, extra=1, can_delete=True)
        formset = DetalleFormSet()
        return render(request, self.template_name, {'form': form, 'formset': formset})

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        DetalleFormSet = inlineformset_factory(Compra, DetalleCompra, form=DetalleCompraForm, extra=1, can_delete=True)
        formset = DetalleFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                self.object = form.save(commit=False)
                self.object.usuario = request.user
                formset.instance = self.object

                total = 0
                detalle_data = []
                producto_cantidades = defaultdict(lambda: Decimal("0"))

                for detalle in formset:
                    if not detalle.cleaned_data or detalle.cleaned_data.get("DELETE"):
                        continue
                    producto = detalle.cleaned_data['producto']
                    cantidad = detalle.cleaned_data['cantidad']
                    precio = detalle.cleaned_data['precio_unitario']
                    detalle_data.append((detalle, producto, cantidad, precio))
                    producto_cantidades[producto.id] += cantidad
                    total += cantidad * precio

                productos = {
                    p.id: p
                    for p in Producto.objects.select_for_update().filter(
                        id__in=producto_cantidades.keys()
                    )
                }

                self.object.total = total
                self.object.save()
                formset.save()

                for prod_id, cantidad_total in producto_cantidades.items():
                    producto = productos[prod_id]
                    producto.stock_actual += cantidad_total
                    producto.save()

                for _, producto, cantidad, _ in detalle_data:
                    MovimientoInventario.objects.create(
                        producto=producto,
                        tipo='entrada',
                        cantidad=cantidad,
                        motivo='Compra',
                        usuario=request.user,
                        operacion_tipo=MovimientoInventario.OPERACION_COMPRA,
                        compra=self.object,
                    )

                messages.success(request, "Compra registrada correctamente.")
                return redirect(self.success_url)

        return render(request, self.template_name, {'form': form, 'formset': formset})

@group_decorator("compras")  
class CompraListView(ListView):
    model = Compra
    template_name = 'core/compra_list.html'
    context_object_name = 'compras'
    ordering = ['-fecha']

@group_decorator("finanzas")
class BalanceView(FormView):
    template_name = 'core/balance.html'
    form_class = BalanceForm
    success_url = reverse_lazy('balance')

    def form_valid(self, form):
        mes = int(form.cleaned_data['mes'])
        anio = int(form.cleaned_data['anio'])

        balance_existente = Balance.objects.filter(mes=mes, anio=anio).first()
        ventas = Venta.objects.filter(fecha__month=mes, fecha__year=anio)
        compras = Compra.objects.filter(fecha__month=mes, fecha__year=anio)

        if balance_existente and balance_existente.cerrado:
            total_ingresos = balance_existente.total_ingresos
            total_egresos = balance_existente.total_egresos
            utilidad = balance_existente.utilidad
            ingresos_operativos = balance_existente.ingresos_operativos
            costos_variables = balance_existente.costos_variables
            costos_fijos = balance_existente.costos_fijos
            gastos_financieros = balance_existente.gastos_financieros
            utilidad_operativa = balance_existente.utilidad_operativa
            utilidad_neta_real = balance_existente.utilidad_neta_real
        else:
            calculo = calcular_balance_mensual(int(mes), int(anio))
            total_ingresos = calculo.total_ingresos
            total_egresos = calculo.total_egresos
            utilidad = calculo.utilidad
            ingresos_operativos = calculo.ingresos_operativos
            costos_variables = calculo.costos_variables
            costos_fijos = calculo.costos_fijos
            gastos_financieros = calculo.gastos_financieros
            utilidad_operativa = calculo.utilidad_operativa
            utilidad_neta_real = calculo.utilidad_neta_real

            Balance.objects.update_or_create(
                mes=mes,
                anio=anio,
                defaults={
                    'total_ingresos': total_ingresos,
                    'total_egresos': total_egresos,
                    'utilidad': utilidad,
                    'ingresos_operativos': ingresos_operativos,
                    'costos_variables': costos_variables,
                    'costos_fijos': costos_fijos,
                    'gastos_financieros': gastos_financieros,
                    'utilidad_operativa': utilidad_operativa,
                    'utilidad_neta_real': utilidad_neta_real,
                }
            )

        context = self.get_context_data(form=form, ventas=ventas, compras=compras,
                                        total_ingresos=total_ingresos,
                                        total_egresos=total_egresos,
                                        utilidad=utilidad,
                                        ingresos_operativos=ingresos_operativos,
                                        costos_variables=costos_variables,
                                        costos_fijos=costos_fijos,
                                        gastos_financieros=gastos_financieros,
                                        utilidad_operativa=utilidad_operativa,
                                        utilidad_neta_real=utilidad_neta_real,
                                        mes=mes, anio=anio)
        return self.render_to_response(context)

@login_decorador   
class MovimientoInventarioListView(ListView):
    model = MovimientoInventario
    template_name = 'core/movimiento_list.html'
    context_object_name = 'movimientos'
    ordering = ['-fecha']

@group_decorator("finanzas")
def exportar_balance_excel(request):
    # Obtener último balance generado
    balance = Balance.objects.order_by('-anio', '-mes').first()

    if not balance:
        return HttpResponse("No hay datos de balance para exportar.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Mensual"

    ws.append(["Balance del Mes", f"{balance.mes:02d}/{balance.anio}"])
    ws.append([])
    ws.append(["Total Ingresos (COP)", balance.total_ingresos])
    ws.append(["Total Egresos (COP)", balance.total_egresos])
    ws.append(["Utilidad (COP)", balance.utilidad])
    ws.append([])
    ws.append(["Ingresos Operativos (COP)", balance.ingresos_operativos])
    ws.append(["Costos Variables (COP)", balance.costos_variables])
    ws.append(["Costos Fijos (COP)", balance.costos_fijos])
    ws.append(["Gastos Financieros (COP)", balance.gastos_financieros])
    ws.append(["Utilidad Operativa (COP)", balance.utilidad_operativa])
    ws.append(["Utilidad Neta Real (COP)", balance.utilidad_neta_real])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=balance_{balance.mes}_{balance.anio}.xlsx'
    wb.save(response)
    return response

@group_decorator("finanzas")
def exportar_balance_pdf(request):
    balance = Balance.objects.order_by('-anio', '-mes').first()

    if not balance:
        return HttpResponse("No hay datos de balance para exportar.")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=balance_{balance.mes}_{balance.anio}.pdf'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 16)
    p.drawString(200, 800, "Balance Mensual")

    p.setFont("Helvetica", 12)
    p.drawString(50, 760, f"Mes/Año: {balance.mes:02d}/{balance.anio}")
    p.drawString(50, 730, f"Total Ingresos: $ {balance.total_ingresos}")
    p.drawString(50, 710, f"Total Egresos: $ {balance.total_egresos}")
    p.drawString(50, 690, f"Utilidad: $ {balance.utilidad}")
    p.drawString(50, 670, f"Ingresos Operativos: $ {balance.ingresos_operativos}")
    p.drawString(50, 650, f"Costos Variables: $ {balance.costos_variables}")
    p.drawString(50, 630, f"Costos Fijos: $ {balance.costos_fijos}")
    p.drawString(50, 610, f"Gastos Financieros: $ {balance.gastos_financieros}")
    p.drawString(50, 590, f"Utilidad Operativa: $ {balance.utilidad_operativa}")
    p.drawString(50, 570, f"Utilidad Neta Real: $ {balance.utilidad_neta_real}")

    p.showPage()
    p.save()
    return response

@login_decorador
class CargarProductosView(View):
    template_name = 'core/cargar_productos.html'

    def get(self, request):
        vista_previa = request.session.pop('vista_previa_productos', None)
        return render(request, self.template_name, {'vista_previa': vista_previa})

    def post(self, request):
        if 'confirmar' in request.POST:
            # Paso 2: Confirmar importación desde sesión
            datos = request.session.get('vista_previa_productos', [])
            creados, errores = 0, 0
            errores_detalle = []
            for fila in datos:
                if fila['estado'] != 'nuevo':
                    continue
                try:
                    categoria_nombre, familia = obtener_categoria_importacion(fila['categoria'])
                    categoria, created = Categoria.objects.get_or_create(
                        nombre_categoria=categoria_nombre,
                        defaults={"familia": familia},
                    )
                    if not created and categoria.familia_id != familia.id:
                        categoria.familia = familia
                        categoria.save(update_fields=["familia"])
                    unidad_media_id = fila.get("unidad_media_id")
                    if not unidad_media_id:
                        raise ValueError(fila.get("unidad_media_error") or "Unidad de medida inválida.")
                    producto = Producto(
                        codigo=fila['codigo'],
                        nombre=fila['nombre'],
                        tipo=fila['tipo'],
                        precio=fila['precio'],
                        stock_actual=fila['stock_actual'],
                        stock_minimo=fila['stock_minimo'],
                        unidad_media_id=unidad_media_id,
                        categoria=categoria,
                        familia=familia,
                    )
                    producto.full_clean()
                    producto.save()
                    creados += 1
                except ValidationError as exc:
                    errores += 1
                    fila_excel = fila.get("fila_excel", "desconocida")
                    if hasattr(exc, "message_dict"):
                        detalles = []
                        for campo, mensajes in exc.message_dict.items():
                            detalles.append(f"{campo}: {', '.join(mensajes)}")
                        mensaje = "; ".join(detalles)
                    else:
                        mensaje = str(exc)
                    errores_detalle.append(f"Fila {fila_excel}: {mensaje}")
                except IntegrityError:
                    errores += 1
                    fila_excel = fila.get("fila_excel", "desconocida")
                    errores_detalle.append(f"Fila {fila_excel}: Código duplicado o conflicto de integridad.")
                except ValueError as exc:
                    errores += 1
                    fila_excel = fila.get("fila_excel", "desconocida")
                    errores_detalle.append(f"Fila {fila_excel}: {exc}")
                except (TypeError, DatabaseError) as exc:
                    errores += 1
                    fila_excel = fila.get("fila_excel", "desconocida")
                    errores_detalle.append(f"Fila {fila_excel}: {exc}")
            if errores_detalle:
                messages.error(request, "Errores en importación:\n" + "\n".join(errores_detalle))
            messages.success(request, f"Se importaron {creados} productos. Errores: {errores}.")
            return redirect('producto_list')

        # Paso 1: Subir archivo y previsualizar
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, "Debe seleccionar un archivo Excel válido.")
            return redirect('cargar_productos')

        path = default_storage.save('tmp/' + archivo.name, archivo)
        wb = load_workbook(default_storage.path(path))
        hoja = wb.active

        vista_previa = []
        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            try:
                codigo, nombre, tipo, precio, stock_actual, stock_minimo, unidad_media, categoria = fila
                estado = 'nuevo'
                error_msg = ''
                if not codigo or not nombre or not precio:
                    estado = 'invalido'
                    error_msg = 'Faltan campos obligatorios.'
                elif Producto.objects.filter(codigo=codigo).exists():
                    estado = 'duplicado'
                    error_msg = 'Código duplicado.'
                unidad, unidad_error = resolver_unidad_media(unidad_media)
                if estado == 'nuevo' and unidad_error:
                    estado = 'invalido'
                    error_msg = unidad_error
                categoria_nombre = categoria or settings.IMPORT_DEFAULT_CATEGORY_NAME
                vista_previa.append({
                    'fila_excel': i,
                    'codigo': codigo or '',
                    'nombre': nombre or '',
                    'tipo': tipo or '',
                    'precio': precio or 0,
                    'stock_actual': stock_actual or 0,
                    'stock_minimo': stock_minimo or settings.IMPORT_DEFAULT_STOCK_MINIMO,
                    'unidad_media': unidad_media or 'unidad',
                    'unidad_media_id': unidad.id if unidad else None,
                    'unidad_media_error': unidad_error,
                    'categoria': categoria_nombre,
                    'estado': estado,
                    'error': error_msg,
                })
            except ValueError:
                vista_previa.append({
                    'fila_excel': i,
                    'codigo': '',
                    'nombre': '',
                    'tipo': '',
                    'precio': 0,
                    'stock_actual': 0,
                    'stock_minimo': 0,
                    'unidad_media': 'unidad',
                    'unidad_media_id': None,
                    'unidad_media_error': "Fila con formato inválido.",
                    'categoria': settings.IMPORT_DEFAULT_CATEGORY_NAME,
                    'estado': 'invalido',
                    'error': 'Fila con formato inválido.',
                })

        request.session['vista_previa_productos'] = vista_previa
        return render(request, self.template_name, {'vista_previa': vista_previa})

@login_decorador   
class DashboardView(TemplateView):
    template_name = 'core/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        hoy = now()
        mes = hoy.month
        anio = hoy.year

        ingresos = Venta.objects.filter(fecha__month=mes, fecha__year=anio).aggregate(total=Sum('total'))['total'] or 0
        egresos = Compra.objects.filter(fecha__month=mes, fecha__year=anio).aggregate(total=Sum('total'))['total'] or 0
        utilidad = ingresos - egresos

        top = DetallesVenta.objects.filter(venta__fecha__month=mes, venta__fecha__year=anio) \
            .values('producto__nombre') \
            .annotate(total_vendido=Sum('cantidad')) \
            .order_by('-total_vendido')

        producto_top = top[0] if top else {'producto__nombre': 'N/A', 'total_vendido': 0}

        ventas_dia_dict = defaultdict(float)
        for venta in Venta.objects.filter(fecha__month=mes, fecha__year=anio):
            dia = venta.fecha.strftime('%d')
            ventas_dia_dict[dia] += float(venta.total)

        compras_dia_dict = defaultdict(float)
        for compra in Compra.objects.filter(fecha__month=mes, fecha__year=anio):
            dia = compra.fecha.strftime('%d')
            compras_dia_dict[dia] += float(compra.total)

        ventas_dia = [{'day': k, 'total': v} for k, v in sorted(ventas_dia_dict.items())]
        compras_dia = [{'day': k, 'total': v} for k, v in sorted(compras_dia_dict.items())]

        # Top 5 productos para gráfica
        top_nombres = [p['producto__nombre'] for p in top[:5]]
        top_cantidades = [p['total_vendido'] for p in top[:5]]

        context.update({
            'mes': mes,
            'anio': anio,
            'ingresos': ingresos,
            'egresos': egresos,
            'utilidad': utilidad,
            'producto_top': producto_top,
            'ventas_dia': ventas_dia,
            'compras_dia': compras_dia,
            'ventas_labels': [v['day'] for v in ventas_dia],
            'ventas_totales': [v['total'] for v in ventas_dia],
            'compras_labels': [c['day'] for c in compras_dia],
            'compras_totales': [c['total'] for c in compras_dia],
            'top_nombres': top_nombres,
            'top_cantidades': top_cantidades
        })
        return context
    
@login_decorador
class ReporteInventarioView(ListView):
    model = Producto
    template_name = 'core/reporte_inventario.html'
    context_object_name = 'productos'

@login_required
def exportar_inventario_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    ws.append(['Codigo', 'Nombre', 'Tipo', 'Stock', 'Stock minimo', 'Unidad', 'Categoria'])
    for p in Producto.objects.all():
        ws.append(
            [
                p.codigo,
                p.nombre,
                p.get_tipo_display(),
                p.stock_actual,
                p.stock_minimo,
                str(p.unidad_media),
                str(p.categoria),
            ]
        )
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventario.xlsx'
    wb.save(response)
    return response

@login_decorador
class MovimientoManualCreateView(CreateView):
    model = MovimientoInventario
    form_class = MovimientoInventarioForm
    template_name = 'core/movimiento_form.html'
    success_url = reverse_lazy('movimiento_list')

    def form_valid(self, form):
        form.instance.usuario = self.request.user if self.request.user.is_authenticated else None
        form.instance.operacion_tipo = MovimientoInventario.OPERACION_AJUSTE
        with transaction.atomic():
            prod = Producto.objects.select_for_update().get(pk=form.instance.producto_id)
            cant = form.cleaned_data["cantidad"]
            tipo = form.cleaned_data["tipo"]
            if tipo != 'entrada' and prod.stock_actual < cant:
                form.add_error('cantidad', 'Stock insuficiente para este producto')
                return self.form_invalid(form)
            response = super().form_valid(form)
            if tipo == 'entrada':
                prod.stock_actual += cant
            else:
                prod.stock_actual -= cant
            prod.save(update_fields=["stock_actual"])
        messages.success(self.request, "Movimiento registrado correctamente.")
        return response
    

class CriticalProductListView(View):
    def get(self, request):
        productos_criticos = Producto.objects.filter(stock_actual__lte=F('stock_minimo')).values(
            'id', 'nombre', 'stock_actual', 'stock_minimo'
        )
        return JsonResponse(list(productos_criticos), safe=False)
    

@login_required
def exportar_perdidas_excel(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    data = calcular_perdidas_devolucion(start, end)

    wb = Workbook()
    ws_cause = wb.active
    ws_cause.title = 'Por causa'
    ws_cause.append(['Motivo', 'Perdida'])
    for causa, total in data['by_cause'].items():
        ws_cause.append([causa, total])

    ws_month = wb.create_sheet('Por mes')
    ws_month.append(['Mes', 'Perdida'])
    for mes, total in data['by_month'].items():
        ws_month.append([mes, total])

    ws_type = wb.create_sheet('Por tipo')
    ws_type.append(['Tipo', 'Perdida'])
    for tipo, total in data['by_type'].items():
        ws_type.append([tipo, total])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=perdidas.xlsx'
    wb.save(response)
    return response
